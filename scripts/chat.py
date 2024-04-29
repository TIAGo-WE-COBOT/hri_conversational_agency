#!/usr/bin/env python3

''' This script is a template to develop your own node implementing a conversational agent. One should override listen and talk methods to define custom behavior.
'''
from emit_sound import EmitSound
from my_timer import MyTimer
from hri_conversational_agency.openai_utils.chatter import OpenAIChatter
from rasa_chat import RasaChatBot
from openpyxl import Workbook, load_workbook
import os
import json
import random
import rospkg
import rospy
import random
from std_msgs.msg import String

class ChatBot():
    def __init__(self):
        self.state = "init" #DEFINE all the states 
        '''
        init
        dummy
        idle
        listen
        talk
        play_media
        exit
        '''

        self.tot_duration = 60 #duration of the effective conversation

        self.dummy_speech = ["domanda 1"] #, \
                             #"domanda 2"]#, \
                             #"domanda 3"]
        self.dummy_question = 0

        self.real_conv = False #to switch betw the dummy and the real conversation
        # Initialize the OpenAI model to generate responses 
        self.ai_chatter = OpenAIChatter()
        self.rasa_chatter = RasaChatBot()
        self.r_sound = EmitSound()

        # Listen to the model response
        # When the model provides a response it triggers the method to let the user add a request
        self.h_req_sub = rospy.Subscriber('r_response',
                                        String,
                                        self.h_listen_cb
                                        )

        # listen to the user input
        # When the user add a request it triggers the model to response to the user input
        # It has to suscribe to the topic used by ros_whisper
        self.r_req_sub = rospy.Subscriber('asr/string',
                                        String,
                                        self.r_listen_cb
                                        )
        
        self.chat_timer_sub = rospy.Subscriber('check/init_timer', 
                                                  String, 
                                                  self.chat_timer_cb)

        self.conv_timer_sub = rospy.Subscriber('check/conv_timer', 
                                                  String, 
                                                  self.end_timer_cb)

        # publisher relative to the robot
        self.r_res_pub = rospy.Publisher('r_response',
                                        String,
                                        queue_size=1
                                        )
        
        self.timer_pb = rospy.Publisher('check/init_timer', 
                                        String, 
                                        queue_size=1
                                        )
        
        self.conv_timer_pb = rospy.Publisher('check/conv_timer',  #forse non serve
                                        String, 
                                        queue_size=1
                                        )
        
        self.play_media_pub = rospy.Publisher('play/media',
                                             String,
                                             queue_size=1
                                             )

        self.slide_col = 0
        self.init_flag = False
        self.dummy_flag = False
        self.idle_flag = False
        self.h_listen_flag = True
        self.r_listen_flag = False
        self.r_talk_flag = False
        self.id_flag = False
        self.chat_init_timer_flag = True
        #self.end_timer_flag = False

        #randomization
        self.exc_n = set()
        self.gen_n = []
        self.mod = ["LLM", "P_LLM", "CHATBOT"]
        self.media = ["M", "V", "AL"]
        self.trials = [[], [], []]

        self.model_response = ""

        self.filename = "BFI_assessment_module_TRIAL2def.xlsx"

    def init_timer(self, duration):
        rospy.Timer(rospy.Duration(duration), self.init_timer_cb, oneshot=True)
        print("init_timer")
        
    def init_timer_cb(self, event):
        self.init_timer_string = String()
        self.init_timer_string.data = "False"
        self.timer_pb.publish(self.init_timer_string)

    def chat_timer_cb(self, timer_chat):
        if(timer_chat.data == "False"):
            self.chat_init_timer_flag = False
            print("Initialization completed")
            rospy.sleep(1)
            self.setup()

    def conv_timer(self, duration): #to count in s the duration of the actual conversation
        rospy.Timer(rospy.Duration(duration), self.conv_timer_cb, oneshot=True)

    def conv_timer_cb(self, event):
        self.conv_timer_string = String()
        self.conv_timer_string.data = "True"
        self.conv_timer_pb.publish(self.conv_timer_string)


    def end_timer_cb(self, end_chat):
        if(end_chat.data == "True"):
            self.ai_chatter.end_timer_flag = True #to be checked to see if the conversation is done
            print("Last interaction")


    def randomization(self, n1, n2, n_max, n_trials, trial1, trial2, trial3):   
        for trial in range(n_trials):
            while len(self.gen_n) < n_max:
                rn = random.randint(n1, n2)
                if(rn not in self.exc_n):
                    if(trial == 0):
                        self.trials[len(self.gen_n)].append(self.mod[(rn)-1])
                        self.gen_n.append(rn)
                        self.exc_n.add(rn)
                    elif(trial == 1):
                        self.trials[len(self.gen_n)].append(self.media[(rn)-1])
                        self.gen_n.append(rn)
                        self.exc_n.add(rn)
                    else:
                        self.trials[len(self.gen_n)].append(rn)
                        self.gen_n.append(rn)
                        self.exc_n.add(rn)
            self.gen_n = []
            self.exc_n = set()
        self.trials = sorted(self.trials, key=lambda x: x[2])
        self.pers_data[str(trial1)] = self.trials[0][0] + "\n" + self.trials[0][1]
        self.pers_data[str(trial2)] = self.trials[1][0] + "\n" + self.trials[1][1]
        self.pers_data[str(trial3)] = self.trials[2][0] + "\n" + self.trials[2][1]
        print(self.trials)
        self.work_space.save(filename = self.filename)

    def setup(self):
        #self.randomization(1, 3, 3, 3)
        if((not(self.init_flag)) and (self.state == "init") and (not(self.chat_init_timer_flag))):
            print("PHASE: SETUP\n")
            try:
                id = int(input("Inserire l'ID del paziente: "))

            except ValueError:
                print("Inserire un ID valido")
            
            rospack = rospkg.RosPack()
            #pkg_root = rospack.get_path('hri_conversational_agency')
            #print(pkg_root)
            # filepath_chat = os.path.join(pkg_root, 
            #                          'scripts', 
            #                          'BFI_assessment_module_TRIAL2copy.xlsx'
            #                          )
        
            # work_space = load_workbook(filename = filepath_chat, data_only=True) #switch with the real path of the excel file
            self.work_space = load_workbook(filename = self.filename, data_only=True)        
            #work_space = load_workbook(filename = "test.xlsx", data_only=True)
            self.pers_data = self.work_space["Preprocessed_Data"]
            col = self.pers_data["B"]
            
            #slide all the rows in the coloum "codice identificativo (ID)"'till finding the inserted ID
            while not self.id_flag:
                if col[self.slide_col].value == id:
                    row = self.pers_data[self.slide_col + 1]
                    self.row = row
                    self.slide_col = 0
                    self.id_flag = True
                else:
                    self.slide_col += 1
                    self.id_flag = False

            if(row[66].value is None or row[66].value == ""):
                self.randomization(1, 3, 3, 3, row[66].coordinate, row[68].coordinate, row[70].coordinate) #randomization generation and writing of the trials on the excel file
            
            if(row[67].value is None or row[67].value == ""):
                self.ai_chatter.curr_mod = str(row[66].value).split("\n")[0] #first modality
                self.rasa_chatter.rasa_curr_mod = self.ai_chatter.curr_mod

                self.ai_chatter.curr_media = str(row[66].value).split("\n")[1] #first media
                self.rasa_chatter.rasa_curr_media = self.ai_chatter.curr_media

                self.ai_chatter.curr_trial = 1
                self.rasa_chatter.rasa_curr_trial = self.ai_chatter.curr_trial

            elif(row[69].value is None or row[69].value == ""):
                self.ai_chatter.curr_mod = str(row[68].value).split("\n")[0] #secont modality
                self.rasa_chatter.rasa_curr_mod = self.ai_chatter.curr_mod

                self.ai_chatter.curr_media = str(row[68].value).split("\n")[1] #second media
                self.rasa_chatter.rasa_curr_media = self.ai_chatter.curr_media

                self.ai_chatter.curr_trial = 2
                self.rasa_chatter.rasa_curr_trial = self.ai_chatter.curr_trial

            elif(row[71].value is None or row[71].value == ""):
                self.ai_chatter.curr_mod = str(row[70].value).split("\n")[0] #third modality
                self.rasa_chatter.rasa_curr_mod = self.ai_chatter.curr_mod

                self.ai_chatter.curr_media = str(row[70].value).split("\n")[1] #third media
                self.rasa_chatter.rasa_curr_media = self.ai_chatter.curr_media

                self.ai_chatter.curr_trial = 3
                self.rasa_chatter.rasa_curr_trial = self.ai_chatter.curr_trial

            else:
                    print("Il paziente ha eseguito tutti i trials")
                    return                          

            #if(self.ai_chatter.curr_mod == "P_LLM"):
                #skip if not P_LLM       
                #print(row[0].value) #row[n] is a cell object, to return the value in cell use row[n].value
                #row[0].value = row[0].value.rstrip(";").replace(";",", ").lower() #to format the string relative to the interests
            self.gender = row[7].value.capitalize()
            self.age = row[8].value
            self.education = row[9].value
            self.job = row[10].value
            self.interests = row[11].value.rstrip(";").rstrip().replace(";",", ").lower().capitalize()
            self.extraversion = row[57].value
            self.agreeableness = row[59].value
            self.conscientiousness = row[61].value
            self.neuroticism = row[63].value
            self.openness = row[65].value
            
            self.init_flag = True
            if(self.ai_chatter.curr_trial==1): #dummy conversation during the first trial only, check if it work
            #self.state = "idle"
                self.state = "dummy"
            #self.idle()
                self.dummy()
            else:
                self.state = "idle"
                self.idle()

    def dummy(self):
        #remove after checking if it work
        if(self.ai_chatter.curr_trial == 1):
            self.pers_data[str(self.row[67].coordinate)] = "DONE"
            print(self.ai_chatter.curr_trial)
            self.work_space.save(filename = self.filename)
            #return
        elif(self.ai_chatter.curr_trial == 2):
            self.pers_data[str(self.row[69].coordinate)] = "DONE"
            print(self.ai_chatter.curr_trial)
            self.work_space.save(filename = self.filename)
            #return
        elif(self.ai_chatter.curr_trial == 3):
            self.pers_data[str(self.row[71].coordinate)] = "DONE"
            print(self.ai_chatter.curr_trial)
            self.work_space.save(filename = self.filename)
            #return

        if self.state == "dummy" and not self.dummy_flag:
            print("PHASE: DUMMY_CONVERSATION\n")
            self.state = "talk"
            self.dummy_flag= True
            self.r_talk("")
            #During the first dummy interction the whisper_node must not publish anything

    def idle(self):
        if self.state == "idle" and not self.idle_flag:
            print("PHASE: IDLE\n")
            self.timer = MyTimer()
            self.conv_timer(self.tot_duration) #in [s]
            self.state = "listen"
            self.idle_flag = True
            
    def h_listen_cb(self, msg):
        print("PHASE: HUMAN_LISTEN\n")
        if(msg.data=="flag" and self.state=="listen"):
            msg.data = ""
        if(self.idle_flag):
            self.h_listen_flag = True
        

    def r_listen_cb(self, msg):
        print("PHASE: ROBOT_LISTEN...waiting for a trascription\n") 
        if not msg.data: #wait 'till string received
            rospy.logwarn('Empty string received, just skipping.')
            return
        if self.state == "listen":
            self.state = "talk"
            if(self.idle_flag or self.dummy_flag):
                self.r_listen_flag = True
            self.r_talk(msg.data)

    
    # Once the model gets an user input, it provides the response
    def r_talk(self, h_prompt):
        print("PHASE: ROBOT_TALK\n")
        if not self.state == "talk":
            raise ValueError("The `r_talk` cb should not be entered when not in state `talk`. How did you get here?!")
        
        if(self.real_conv):
            print("Real conversation")
            #self.ai_chatter.generate_s_prompt(self.gender, self.age, self.education, self.job, self.interests, self.extraversion, self.agreeableness, self.conscientiousness, self.neuroticism, self.openness)
            #if(self.ai_chatter.play_media_flag):
            #print("line 303")
            #self.state = "play_media"
            #self.media_player() ##########################add method media_player
            #ad flag???
            #decidere se generare il prompt con il setup oppure con la generazione della risposta
            #self.ai_chatter.generate_s_prompt(self.gender, self.age, self.education, self.job, self.interests, self.extraversion, self.agreeableness, self.conscientiousness, self.neuroticism, self.openness)
            if(self.ai_chatter.curr_mod == "LLM" or self.ai_chatter.curr_mod == "P_LLM"):
                if(self.idle_flag and self.r_listen_flag and self.h_listen_flag):
                    self.ai_chatter.n_interactions = self.ai_chatter.n_interactions + 1 #
                    self.h_listen_flag = False
                    self.r_listen_flag = False

                self.ai_chatter.generate_s_prompt(self.gender, self.age, self.education, self.job, self.interests, self.extraversion, self.agreeableness, self.conscientiousness, self.neuroticism, self.openness)
                #r_ans = self.ai_chatter.generate_response(self.ai_chatter.n_interactions, self.ai_chatter.s_prompt, h_prompt) #string genereted by the model #
                #self.model_response = r_ans #model
                if(self.ai_chatter.end_timer_flag):
                    if(not self.ai_chatter.find_media_flag):
                        r_ans = "Ciao, come stai?"+ "Canzone 1, canzone 2 oppure canzone 3"
                    else:
                         #Use this line to test the system without wasting tokens
                        r_ans = "2"
                else:
                    r_ans = "Ciao, come stai?"
                self.model_response = r_ans
                self.ai_chatter.conversational_hystory(h_prompt, r_ans)
                self.ai_chatter.log.log_curr_interaction(self.ai_chatter.n_interactions)
                self.ai_chatter.log.log_system_prompt(self.ai_chatter.s_prompt)
                self.ai_chatter.log.log_input(h_prompt)
                self.ai_chatter.log.log_output(r_ans, self.ai_chatter.model)
                self.ai_chatter.log.log_input_tokens(cb.ai_chatter.prompt_tokens)
                self.ai_chatter.log.log_output_tokens(cb.ai_chatter.compl_tokens)
                self.ai_chatter.log.log_tot_tokens(cb.ai_chatter.tot_tokens)

                if(not self.ai_chatter.play_media_flag):        
                    self.r_sound.r_say(r_ans)
                    r_msg = String()
                    r_msg.data = "flag"
                    self.r_res_pub.publish(r_msg)
                    r_msg.data = ""
                    self.state = "listen"
                else:
                    self.state = "play_media"
                    self.media_player() ##########################add method media_player
                 
            elif(self.ai_chatter.curr_mod == "CHATBOT"):
                if(self.idle_flag and self.r_listen_flag and self.h_listen_flag):
                    self.rasa_chatter.rasa_interaction += 1 
                    self.h_listen_flag = False
                    self.r_listen_flag = False
                
                #togliere questo if e mettere if in rasa chatter

                self.rasa_chatter.send_request(h_prompt)
                print(self.rasa_chatter.chatbot_resp)
                ###mettere i log come per llm e pllm, bisogna aggiungere metodi nello script rasa_chat

                ######################################################################################
                if(not self.rasa_chatter.rasa_play_media_flag):    
                    self.r_sound.r_say(self.rasa_chatter.chatbot_resp)
                    chatbot_msg = String()
                    chatbot_msg.data = "flag"
                    self.r_res_pub.publish(chatbot_msg)
                    chatbot_msg.data = ""
                    self.state = "listen"
                else:
                    self.state = "play_media"
                    self.media_player() ##########################add method media_player

        elif(not self.real_conv):
        #dummy conversation during the first trial only, check if it work
            self.r_sound.r_say(self.dummy_speech[self.dummy_question])
            self.dummy_question += 1
            r_msg = String()
            r_msg.data = "flag"
            self.r_res_pub.publish(r_msg)
            r_msg.data = ""
            self.state = "listen"
            if(self.dummy_question > (len(self.dummy_speech)-1)):
                self.real_conv = True
                r_msg.data = ""
                self.r_res_pub.publish(r_msg)
                self.state = "idle"
                self.idle()
        # r_msg = String()
        # r_msg.data = "flag"
        # self.r_res_pub.publish(r_msg)
        # r_msg.data = ""
        # self.state = "listen"

    def media_player(self): #use self.model response
        if(self.state == "send_media"):
            media = String()
            if(self.ai_chatter.curr_media == "M" or self.rasa_chatter.rasa_curr_media == "M"):
                self.r_sound.r_say("Riproduco musica")
                data = 'M'+'%'+'musica da riprodurre'
                media.data = data
                self.play_media_pub.publish(media)
            elif(self.ai_chatter.curr_media == "V" or self.rasa_chatter.rasa_curr_media == "V"):
                self.r_sound.r_say("Mostro un video")
                data = 'V'+'%'+'video da riprodurre'
                media.data = data
                self.play_media_pub.publish(media)
            elif(self.ai_chatter.curr_media == "AL" or self.rasa_chatter.rasa_curr_media == "AL"):
                self.r_sound.r_say("Riproduco un audiolibro")
                data = 'AL'+'%'+'audiolibro da riprodurre'
                media.data = data
                self.play_media_pub.publish(media)

            #rospy.sleep(1) #wait for some second
            print("TRIAL DONE")
            self.state = "exit"
        #check the media that has to be played

if __name__ == "__main__":
    rospy.init_node('chatbot')
    cb = ChatBot()
    def on_shutdown():
        cb.ai_chatter.log.log_n_interactions(cb.ai_chatter.n_interactions) #
        duration = cb.timer.elapsed_time()
        cb.ai_chatter.log.log_duration(duration)
        cb.ai_chatter.log.log_tot_tokens(cb.ai_chatter.tot_tokens)
        cb.ai_chatter.log.log_close
        print(json.dumps(cb.ai_chatter.conversation, indent=2))

    rospy.on_shutdown(on_shutdown)
    rate = rospy.Rate(10)
    try:
        #cb.idle()
        cb.init_timer(3)
        rospy.spin()
    except KeyboardInterrupt:
        rospy.loginfo('Shutting down on user request.')
            
